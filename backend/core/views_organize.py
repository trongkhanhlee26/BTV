from django.shortcuts import render, redirect
from django.contrib import messages
from django.views.decorators.http import require_http_methods
from django.db.models import Prefetch
from .models import CuocThi, VongThi, BaiThi, BaiThiTimeRule, BaiThiTemplateSection, BaiThiTemplateItem



@require_http_methods(["GET", "POST"])
def organize_view(request):
    if request.method == "POST":
        action = request.POST.get("action")
        try:
            if action == "create_ct":
                ten = (request.POST.get("tenCuocThi") or "").strip()
                trang_thai = request.POST.get("trangThai") == "on"
                if not ten:
                    messages.error(request, "Vui lòng nhập tên cuộc thi.")
                else:
                    ct = CuocThi(tenCuocThi=ten, trangThai=trang_thai)
                    ct.save()  # auto-gen ct.ma
                    messages.success(request, f"Tạo cuộc thi {ct.ma} thành công.")
                return redirect(request.path)

            if action == "create_vt":
                ct_id = request.POST.get("cuocThi_id")
                ten_vt = (request.POST.get("tenVongThi") or "").strip()
                if not ct_id or not ten_vt:
                    messages.error(request, "Vui lòng chọn cuộc thi và nhập tên vòng thi.")
                else:
                    cuoc_thi = CuocThi.objects.get(id=ct_id)
                    vt = VongThi(tenVongThi=ten_vt, cuocThi=cuoc_thi)
                    vt.save()  # auto-gen vt.ma
                    messages.success(request, f"Tạo vòng thi {vt.ma} trong {cuoc_thi.ma} thành công.")
                return redirect(request.path)

            if action == "create_bt":
                vt_id = request.POST.get("vongThi_id")
                ten_bt = (request.POST.get("tenBaiThi") or "").strip()
                method = request.POST.get("phuongThucCham") or "POINTS"
                max_diem = request.POST.get("cachChamDiem")
                # POINTS thì phải có max_diem; TIME/TEMPLATE thì cho phép max_diem=0 tạm thời
                if not vt_id or not ten_bt:
                    messages.error(request, "Vui lòng chọn vòng thi và nhập tên bài thi.")
                elif method == "POINTS" and not max_diem:
                    messages.error(request, "Vui lòng nhập điểm tối đa cho phương thức thang điểm.")
                else:
                    vong_thi = VongThi.objects.get(id=vt_id)
                    bt = BaiThi(
                        tenBaiThi=ten_bt,
                        vongThi=vong_thi,
                        phuongThucCham=method,
                        cachChamDiem=int(max_diem) if (method == "POINTS" and max_diem) else 0,
                    )
                    bt.save()  # auto-gen bt.ma
                    messages.success(request, f"Tạo bài thi {bt.ma} trong {vong_thi.ma} thành công.")
                return redirect(request.path)
            # NEW: cấu hình thang thời gian
            if action == "config_time_rules":
                import json
                btid = request.POST.get("baiThi_id")
                

                raw = request.POST.get("time_rules_json") or "[]"
                try:
                    bt = BaiThi.objects.get(id=btid)
                except BaiThi.DoesNotExist:
                    messages.error(request, "Bài thi không tồn tại.")
                    return redirect(request.path)
                if bt.phuongThucCham != "TIME":
                    messages.error(request, "Bài thi này không dùng phương thức chấm theo thời gian.")
                    return redirect(request.path)
                try:
                    rows = json.loads(raw)
                except Exception:
                    messages.error(request, "Dữ liệu cấu hình không hợp lệ.")
                    return redirect(request.path)
                # validate sơ bộ & lưu
                cleaned = []
                for r in rows:
                    try:
                        s = int(r.get("start", 0))
                        e = int(r.get("end", 0))
                        sc = int(r.get("score", 0))
                    except Exception:
                        continue
                    if s < 0 or e < 0 or e < s:
                        continue
                    cleaned.append((s, e, sc))
                from django.db import transaction
                with transaction.atomic():
                    bt.time_rules.all().delete()
                    BaiThiTimeRule.objects.bulk_create([
                        BaiThiTimeRule(baiThi=bt, start_seconds=s, end_seconds=e, score=sc)
                        for (s, e, sc) in cleaned
                    ])
                messages.success(request, f"Đã lưu {len(cleaned)} dòng thang thời gian cho {bt.ma}.")
                return redirect(request.path)

                        # NEW: cấu hình thang thời gian
            if action == "config_time_rules":
                import json
                btid = request.POST.get("baiThi_id")
                messages.success(request, f"Đã lưu {len(cleaned)} dòng thang thời gian cho {bt.ma}.")
                return redirect(request.path)

            # NEW: cấu hình mẫu chấm (import Excel) — chạy tại /organize/
            if action == "config_template_upload":
                btid = request.POST.get("baiThi_id")
                f = request.FILES.get("template_file")
                if not btid or not f:
                    messages.error(request, "Thiếu bài thi hoặc tệp Excel.")
                    return redirect(request.path)

                try:
                    bt = BaiThi.objects.get(id=btid)
                except BaiThi.DoesNotExist:
                    messages.error(request, "Bài thi không tồn tại.")
                    return redirect(request.path)

                if bt.phuongThucCham != "TEMPLATE":
                    messages.error(request, "Bài thi này không dùng phương thức chấm theo mẫu.")
                    return redirect(request.path)

                # đọc Excel
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(f, data_only=True)
                    ws = wb[wb.sheetnames[0]]
                except Exception:
                    messages.error(request, "Không đọc được file Excel. Vui lòng dùng .xlsx hợp lệ.")
                    return redirect(request.path)

                # ===== NEW: Quét header linh hoạt theo mẫu "Danh Mục 1 | Danh Mục 2 | Điểm | (Điểm Chấm)" =====
                header_row_idx = None
                col_section = col_item = col_max = None

                def norm(v):
                    return str(v).strip().lower() if v is not None else ""

                # 1) Tìm hàng header trong 30 hàng đầu
                for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
                    texts = [norm(c) for c in row]
                    # dò theo từ khoá Việt của mẫu
                    for cidx, t in enumerate(texts):
                        if t in ("danh mục 1", "mục lớn", "section"):
                            col_section = cidx
                        elif t in ("danh mục 2", "mục nhỏ", "item", "nội dung"):
                            col_item = cidx
                        elif ("điểm" in t) and ("chấm" not in t):
                            # lấy cột "Điểm" (tối đa), nhưng bỏ "Điểm Chấm"
                            col_max = cidx
                    if col_section is not None and col_item is not None and col_max is not None:
                        header_row_idx = ridx
                        break

                if header_row_idx is None:
                    messages.error(request, "Không tìm thấy dòng tiêu đề (Danh Mục 1 / Danh Mục 2 / Điểm).")
                    return redirect(request.path)

                # 2) (Tuỳ chọn) Bắt 'Tổng điểm tối đa' ở khu vực trên cùng (đơn giản, không dùng cú pháp đặc biệt)
                total_max = None
                for ridx, row in enumerate(
                        ws.iter_rows(min_row=1, max_row=min(header_row_idx, 10), min_col=1, max_col=20, values_only=False),
                        start=1):
                    for cidx, cell in enumerate(row, start=1):  # cidx là chỉ số 1-based
                        t = norm(cell.value)
                        if "tổng điểm tối đa" in t:
                            # Tìm số ở tối đa 5 ô bên phải
                            for off in range(1, 6):
                                # chuyển về index 0-based cho list `row`
                                idx = (cidx - 1) + off
                                if idx >= len(row):
                                    break
                                val = row[idx].value
                                if isinstance(val, (int, float)):
                                    total_max = int(val)
                                    break
                            break
                    if total_max is not None:
                        break

                # 3) Gom dữ liệu từ hàng ngay sau header
                rows = []
                for r in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                    v_section = norm(r[col_section]) if len(r) > col_section else ""
                    v_item    = norm(r[col_item])    if len(r) > col_item    else ""
                    v_max_raw = r[col_max]           if len(r) > col_max     else None

                    # Dừng khi qua nhiều dòng trống liên tiếp (không cả section & item)
                    if not v_section and not v_item and v_max_raw in (None, ""):
                        # cho phép vài dòng trống xen kẽ, nhưng nếu muốn dừng hẳn thì dùng break
                        continue

                    # ép kiểu điểm tối đa
                    mx = 0
                    if isinstance(v_max_raw, (int, float)):
                        mx = int(v_max_raw)
                    else:
                        try:
                            mx = int(str(v_max_raw).strip()) if v_max_raw not in (None, "") else 0
                        except Exception:
                            mx = 0

                    # bỏ các dòng placeholder như "nhập điểm" ở cột "Điểm Chấm" (không liên quan vì ta không đọc cột đó)
                    if not v_section and not v_item:
                        continue

                    # NOTE: ghi nhận cả khi section có mà item trống (mục lớn không có lá)
                    # nhưng để mô hình hiện tại, ta chỉ tạo item khi có cả 2
                    if v_section and v_item:
                        rows.append((v_section, v_item, mx, ""))

                if not rows:
                    messages.error(request, "Không tìm thấy dòng dữ liệu nào dưới tiêu đề (Danh Mục 1/2, Điểm).")
                    return redirect(request.path)

                # 4) (Tuỳ chọn) cập nhật điểm tối đa tổng nếu bạn có trường (bỏ qua nếu không có)
                try:
                    if total_max is not None and hasattr(bt, "diemToiDa"):
                        bt.diemToiDa = int(total_max)
                        bt.save(update_fields=["diemToiDa"])
                except Exception:
                    pass
                # ===== END NEW =====


                # lưu vào DB
                from django.db import transaction
                with transaction.atomic():
                    bt.template_sections.all().delete()

                    section_order, section_index = [], {}
                    for (sect, _, _, _) in rows:
                        if sect not in section_index:
                            section_index[sect] = len(section_order) + 1
                            section_order.append(sect)

                    BaiThiTemplateSection.objects.bulk_create([
                        BaiThiTemplateSection(baiThi=bt, stt=section_index[name], title=name)
                        for name in section_order
                    ])

                    created_sections = {
                        s.title: s for s in BaiThiTemplateSection.objects.filter(baiThi=bt)
                    }
                    counters = {name: 0 for name in section_order}
                    items = []
                    for (sect, item, mx, note) in rows:
                        counters[sect] += 1
                        items.append(BaiThiTemplateItem(
                            section=created_sections[sect],
                            stt=counters[sect],
                            content=item,
                            max_score=mx,
                            note=(note or None),
                        ))
                    BaiThiTemplateItem.objects.bulk_create(items)

                messages.success(request, f"Đã import {len(section_order)} mục lớn và {len(items)} mục nhỏ cho {bt.ma}.")
                return redirect(request.path)

            messages.error(request, "Hành động không hợp lệ.")
            return redirect(request.path)

            

        except CuocThi.DoesNotExist:
            messages.error(request, "Cuộc thi không tồn tại.")
            return redirect(request.path)
        except VongThi.DoesNotExist:
            messages.error(request, "Vòng thi không tồn tại.")
            return redirect(request.path)
        except ValueError:
            messages.error(request, "Giá trị điểm tối đa không hợp lệ.")
            return redirect(request.path)

    # GET: hiển thị danh sách CT → VT → BT
    cuoc_this = (
        CuocThi.objects
        .prefetch_related(
            Prefetch(
                "vong_thi",
                queryset=VongThi.objects.prefetch_related(
                    "bai_thi__time_rules",
                    "bai_thi__template_sections__items",  # NEW
                )
            )
        )
        .order_by("-id")
    )

    return render(request, "organize/index.html", {"cuoc_this": cuoc_this})


@require_http_methods(["GET", "POST"])
def competition_list_view(request):
    if request.method == "POST":
        action = request.POST.get("action")
        try:
            if action == "create":
                ten = (request.POST.get("tenCuocThi") or "").strip()
                trang_thai = request.POST.get("trangThai") == "on"
                if not ten:
                    messages.error(request, "Vui lòng nhập tên cuộc thi.")
                else:
                    ct = CuocThi(tenCuocThi=ten, trangThai=trang_thai)
                    ct.save()
                    messages.success(request, f"Đã tạo {ct.ma}.")
                return redirect(request.path)

            if action == "update":
                ct_id = request.POST.get("id")
                ten = (request.POST.get("tenCuocThi") or "").strip()
                trang_thai = request.POST.get("trangThai") == "on"
                ct = CuocThi.objects.get(id=ct_id)
                if not ten:
                    messages.error(request, "Tên cuộc thi không được rỗng.")
                else:
                    ct.tenCuocThi = ten
                    ct.trangThai = trang_thai
                    ct.save(update_fields=["tenCuocThi", "trangThai"])
                    messages.success(request, f"Đã cập nhật {ct.ma}.")
                return redirect(request.path)

            if action == "delete":
                ct_id = request.POST.get("id")
                ct = CuocThi.objects.get(id=ct_id)
                code = ct.ma
                ct.delete()
                messages.success(request, f"Đã xoá {code}.")
                return redirect(request.path)
                        # NEW: cấu hình mẫu chấm (import Excel)
            

            messages.error(request, "Hành động không hợp lệ.")
            return redirect(request.path)
        except CuocThi.DoesNotExist:
            messages.error(request, "Cuộc thi không tồn tại.")
            return redirect(request.path)

    # GET
    items = CuocThi.objects.order_by("-id")
    return render(request, "organize/competitions.html", {"items": items})