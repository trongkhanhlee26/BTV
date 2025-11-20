# core/views_export.py
from __future__ import annotations
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404
from django.db.models import Avg
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side  # <- thêm Border, Side
from .models import CuocThi, VongThi, BaiThi, ThiSinh, PhieuChamDiem


# --- helpers cho thời gian ---
def _pick_time_value(obj):
    """
    Trích xuất thời gian (giây) từ một record PhieuChamDiem.
    Hỗ trợ linh hoạt nhiều tên field khác nhau.
    Trả về int(giây) hoặc None nếu không có.
    """
    CANDIDATES = ["thoiGian", "thoiGianGiay", "time_seconds", "time", "duration", "tongThoiGian"]
    for k in CANDIDATES:
        if hasattr(obj, k):
            v = getattr(obj, k)
            if v is None: 
                continue
            try:
                # chấp nhận float/decimal → ép int giây
                return int(round(float(v)))
            except Exception:
                pass
    return None

def _fmt_mmss(seconds: int | None) -> str:
    if seconds is None:
        return ""
    m, s = divmod(max(0, int(seconds)), 60)
    return f"{m:02d}:{s:02d}"

def _score_type(bt) -> str:
    v = getattr(bt, "phuongThucCham", None)
    if v is None:
        return "POINTS"
    s = str(v).strip().upper()
    if s in {"TIME", "2"}:
        return "TIME"
    if s in {"TEMPLATE", "1"}:
        return "TEMPLATE"
    return "POINTS"

def _build_columns(ct: CuocThi):
    vong_ids = VongThi.objects.filter(cuocThi=ct).values_list("id", flat=True)
    bai_qs = (
        BaiThi.objects
        .filter(vongThi_id__in=vong_ids)
        .select_related("vongThi")
        .prefetch_related("time_rules", "template_sections__items")
        .order_by("vongThi_id", "id")
    )

    cols = []
    for b in bai_qs:
        # tính max điểm để hiện trên header (giữ nguyên logic cũ)
        if _score_type(b) == "TIME":
            rules = list(b.time_rules.all()) if hasattr(b, "time_rules") else []
            b_max = max([r.score for r in rules], default=0)
        elif _score_type(b) == "TEMPLATE":
            b_max = sum(i.max_score for s in b.template_sections.all() for i in s.items.all())
        else:
            b_max = b.cachChamDiem

        # 1) Cột điểm (giữ tiêu đề 2 dòng để JS nhận diện là cột điểm)
        cols.append({
            "id": b.id,
            "code": b.ma,
            "kind": "score",
            "title": f"{b.vongThi.tenVongThi}\n{b.tenBaiThi}",
            "max": b_max,
        })
        # 2) Cột thời gian đi kèm (đặt ngay sau cột điểm)
        cols.append({
            "id": b.id,
            "code": b.ma,
            "kind": "time",
            # tiêu đề rõ ràng: “Thời gian (BTxxx)” – có thể rút gọn nếu muốn
            "title": f"Thời gian",
            "max": None,
        })

    titles = [c["title"] for c in cols]   # chỉ tiêu đề phần bài thi
    return cols, titles


def _flatten(ct: CuocThi):
    cols_meta, titles_per_exam = _build_columns(ct)

    # Info columns bạn đang dùng
    info_titles = ['Đơn vị', 'Chi nhánh', 'Vùng', 'Nhóm', 'Email']

    # Xây tiêu đề tổng cộng: 3 cột info cơ bản + 5 info + (cặp Điểm/Thời gian)* + Tổng
    columns = ['STT', 'Mã NV', 'Họ tên'] + info_titles + titles_per_exam + ['Tổng']


    # Map điểm trung bình theo (maNV, baiThi_id)
    score_qs = (
        PhieuChamDiem.objects
        .filter(cuocThi=ct)
        .values("thiSinh__maNV", "baiThi_id")
        .annotate(avg=Avg("diem"))
    )
    score_map = {(r["thiSinh__maNV"], r["baiThi_id"]): (float(r["avg"]) if r["avg"] is not None else "") for r in score_qs}

    # Lấy tất cả phiếu để tự bóc thời gian (vì tên field thời gian có thể khác nhau)
    all_phieu = list(PhieuChamDiem.objects.filter(cuocThi=ct).select_related("thiSinh", "baiThi"))
    time_map = {}  # (maNV, baiThi_id) -> một giá trị thời gian (ưu tiên nhỏ nhất nếu có nhiều lần)
    for p in all_phieu:
        key = (getattr(p.thiSinh, "maNV", None), getattr(p.baiThi, "id", None))
        if key[0] is None or key[1] is None:
            continue
        t = _pick_time_value(p)  # giây hoặc None
        if t is None:
            continue
        # nếu có nhiều bản ghi: lấy MIN (thường là tốt nhất), bạn có thể đổi sang MAX hay AVG tùy nghiệp vụ
        cur = time_map.get(key)
        if (cur is None) or (t < cur):
            time_map[key] = t

    ts_qs = ThiSinh.objects.filter(cuocThi=ct).order_by("maNV").distinct()
    def _sv(x): return "" if x is None else str(x)

    rows = []
    for idx, ts in enumerate(ts_qs, start=1):
        row = [
            idx,
            _sv(getattr(ts, "maNV", "")),
            _sv(getattr(ts, "hoTen", "")),
            _sv(getattr(ts, "donVi", "")),
            _sv(getattr(ts, "chiNhanh", "")),
            _sv(getattr(ts, "vung", "")),
            _sv(getattr(ts, "nhom", "")),
            _sv(getattr(ts, "email", "")),
        ]

        total = 0.0

        # chỉ lấy mỗi bài thi 1 lần (theo các cột 'score')
        bt_ids_in_order = [c["id"] for c in cols_meta if c.get("kind") == "score"]

        for bt_id in bt_ids_in_order:
            # 1) điểm
            sc = score_map.get((ts.maNV, bt_id), "")
            row.append(sc)
            if isinstance(sc, (int, float)):
                total += float(sc)

            # 2) thời gian (mm:ss) – để trống nếu không có
            tm_seconds = time_map.get((ts.maNV, bt_id))
            row.append(_fmt_mmss(tm_seconds))

        row.append(total)
        rows.append(row)

    return columns, rows


# (Giữ export_page như cũ)

# (Tuỳ ý: có thể xoá export_csv và route của nó)
# def export_csv(...):  # <-- BỎ KHI KHÔNG DÙNG NỮA
#     ...

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
def export_xlsx(request):
    ct_id = request.GET.get("ct")
    ct = get_object_or_404(CuocThi, id=ct_id)

    use_visible = (request.method == "POST")
    if use_visible:
        # Nhận payload từ frontend
        import json
        try:
            payload = json.loads(request.body.decode("utf-8"))
            columns = payload.get("columns") or []
            rows = payload.get("rows") or []
            kinds = payload.get("col_kinds") or ["info"] * len(columns)
        except Exception:
            # fallback sang full nếu payload lỗi
            columns, rows = _flatten(ct)
            kinds = ["info"] * len(columns)
    else:
        columns, rows = _flatten(ct)

        info_count = 3 + 5
        kinds = ["info"] * len(columns)

        j = info_count
        # duyệt tới cột áp chót (cột cuối là "Tổng")
        while j < len(columns) - 1:
            # cặp 1: điểm
            kinds[j] = "score"
            j += 1
            if j < len(columns) - 1:
                # cặp 2: thời gian
                kinds[j] = "time"
                j += 1
        kinds[-1] = "score"   # "Tổng"


    wb = Workbook()
    ws = wb.active
    ws.title = f"{ct.ma}"

    # ==== Header: KHÔNG in đậm ====
    ws.append(columns)
    for c in range(1, len(columns)+1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, size=12)  # KHÔNG in đậm
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ==== Body: dùng rows đã có (full hoặc visible) ====
    for r in rows:
        ws.append(r)

    # ==== Tô màu theo CỘT ====
    fill_info  = PatternFill(fill_type="solid", start_color="FFEAF4FF", end_color="FFEAF4FF")  # xanh nhạt
    fill_score = PatternFill(fill_type="solid", start_color="FFFFF5E6", end_color="FFFFF5E6")  # vàng nhạt

    max_row = ws.max_row
    max_col = ws.max_column
    for j in range(1, max_col+1):
        kind = kinds[j-1] if (j-1) < len(kinds) else "info"
        fill = fill_score if kind == "score" else fill_info
        for i in range(1, max_row+1):
            ws.cell(row=i, column=j).fill = fill


    # ... sau khi append header + body và tô màu, tính sẵn:
    max_row = ws.max_row
    max_col = ws.max_column

    # ==== Border mảnh cho toàn bộ ô ====
    thin = Side(style="thin", color="FF000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col), start=1):
        for cell in row:
            cell.border = border_all
            # Font: header (r1) 12pt, body 11pt, không đậm (đúng yêu cầu “không in đậm”)
            if r_idx == 1:
                cell.font = Font(name="Times new roman", size=12, bold=True)
            else:
                cell.font = Font(name="Times new roman", size=11, bold=False)
            # Alignment
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=True,
                horizontal=cell.alignment.horizontal if cell.alignment else "left"
            )

    # (giữ nguyên) Freeze 3 cột + 1 hàng tiêu đề
    ws.freeze_panes = "E2"

    # ==== Auto width (rộng hơn một chút) ====
    for i, col in enumerate(columns, start=1):
        maxlen = len(str(col)) if col is not None else 0
        for r in rows:
            v = r[i-1] if i-1 < len(r) else ""
            l = len(str(v)) if v is not None else 0
            if l > maxlen:
                maxlen = l
        # padding rộng hơn: +4, tối thiểu 12, tối đa 60
        ws.column_dimensions[get_column_letter(i)].width = max(12, min(maxlen + 4, 60))


    # Giữ freeze panes cũ (3 cột trái + 1 hàng tiêu đề)
    ws.freeze_panes = "E2"

    bio = BytesIO(); wb.save(bio); bio.seek(0)
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    # Tên file: nếu POST visible thì đổi chút cho phân biệt
    fname = f'export_{ct.ma}.xlsx' if not use_visible else f'export_{ct.ma}.xlsx'
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp
def export_page(request):
    ct_id = request.GET.get("ct")
    ct = get_object_or_404(CuocThi, id=ct_id)

    # Chỉ lấy các cuộc thi đang bật
    active_cts = CuocThi.objects.filter(trangThai=True).order_by("ma", "tenCuocThi")

    columns, rows = _flatten(ct)
    return render(request, "export/index.html", {
        "contest": ct,
        "columns": columns,
        "rows": rows,
        "active_cts": active_cts,   # <-- thêm vào context
    })
# --- FINAL EXPORT (Chung Kết) ---
from django.db.models import Avg, Sum
from django.http import JsonResponse
from .models import CuocThi, VongThi, BaiThi, ThiSinh, PhieuChamDiem, BattleVote

def _find_chung_ket():
    """
    Tìm cuộc thi 'Chung Kết' theo nhiều biến thể: 'Chung Kết' / 'Chung Ket' (không dấu).
    Ưu tiên __iexact, fallback bản không dấu thô.
    """
    ct = CuocThi.objects.filter(tenCuocThi__iexact="Chung Kết").first()
    if not ct:
        ct = CuocThi.objects.filter(tenCuocThi__iexact="Chung Ket").first()
    return ct

def _final_columns_and_rows(ct: CuocThi):
    """
    Trả về (columns, rows) cho trang Export Chung Kết:
    - Cột info như export thường: STT, Mã NV, Họ tên, Đơn vị, Chi nhánh, Vùng, Nhóm, Email
    - Cột điểm: Tổng điểm (vòng Chung kết), Đối kháng (sao TB, 1 số thập phân)
    """
    info_titles = ['STT', 'Mã NV', 'Họ tên', 'Đơn vị', 'Chi nhánh', 'Vùng', 'Nhóm', 'Email']
    columns = info_titles + ['Tổng điểm', 'Đối kháng']

    # 1) Xác định Vòng “Chung Kết” (nếu không tìm được thì lấy tất cả vòng của CT này)
    vt_ck = VongThi.objects.filter(cuocThi=ct, tenVongThi__iexact="Chung Kết")
    if not vt_ck.exists():
        vt_ck = VongThi.objects.filter(cuocThi=ct)
    vt_ids = list(vt_ck.values_list("id", flat=True))

    # 2) Các bài thi thuộc vòng CK
    bt_ids = list(BaiThi.objects.filter(vongThi_id__in=vt_ids).values_list("id", flat=True))

    # 3) Map tổng điểm vòng CK cho từng thí sinh
    #    (gộp trung bình theo bài, rồi SUM các bài)
    score_qs = (
        PhieuChamDiem.objects
        .filter(cuocThi=ct, baiThi_id__in=bt_ids)
        .values("thiSinh__maNV", "baiThi_id")
        .annotate(avg=Avg("diem"))
    )
    # tích lũy SUM(avg) theo thí sinh
    total_by_ma = {}
    for r in score_qs:
        ma = r["thiSinh__maNV"]
        total_by_ma[ma] = total_by_ma.get(ma, 0.0) + float(r["avg"] or 0.0)

    # 4) Đối kháng: trung bình sao theo BattleVote cho CT này
    #    BattleVote.entry -> ThiSinhCapThiDau -> pair -> cuocThi
    #    Lấy TB sao theo thiSinh (entry.thiSinh.maNV)
    battle_qs = (
        BattleVote.objects
        .filter(entry__pair__cuocThi=ct)
        .values("entry__thiSinh__maNV")
        .annotate(avg=Avg("stars"))
    )
    stars_by_ma = {r["entry__thiSinh__maNV"]: (float(r["avg"]) if r["avg"] is not None else None)
                   for r in battle_qs}

    # 5) Duyệt thí sinh của CT & build rows
    ts_qs = ThiSinh.objects.filter(cuocThi=ct).order_by("maNV").distinct()
    def _sv(x): return "" if x is None else str(x)

    rows = []
    for idx, ts in enumerate(ts_qs, start=1):
        tong = round(total_by_ma.get(ts.maNV, 0.0), 2)
        sao = stars_by_ma.get(ts.maNV, None)
        sao_fmt = (f"{sao:.1f}" if sao is not None else "")

        row = [
            idx,
            _sv(getattr(ts, "maNV", "")),
            _sv(getattr(ts, "hoTen", "")),
            _sv(getattr(ts, "donVi", "")),
            _sv(getattr(ts, "chiNhanh", "")),
            _sv(getattr(ts, "vung", "")),
            _sv(getattr(ts, "nhom", "")),
            _sv(getattr(ts, "email", "")),
            tong,
            sao_fmt,
        ]
        rows.append(row)

    return columns, rows

def export_final_page(request):
    """
    Trang web Export Chung Kết (bảng Excel-like).
    Cố định CT = 'Chung Kết' (không hiển thị dropdown chọn CT).
    """
    ct = _find_chung_ket()
    if not ct:
        return render(request, "export/index.html", {
            "contest": None,
            "columns": [],
            "rows": [],
            "FROZEN_COUNT": 3,
            "final_mode": True,   # flag cho UI nếu muốn
            "error": "Chưa tạo cuộc thi 'Chung Kết'."
        })

    columns, rows = _final_columns_and_rows(ct)
    return render(request, "export/index.html", {
        "contest": ct,
        "columns": columns,
        "rows": rows,
        "FROZEN_COUNT": 3,
        "final_mode": True
    })

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from io import BytesIO

def export_final_xlsx(request):
    """
    Xuất XLSX cho Chung Kết (giống export-xlsx nhưng chỉ 2 cột điểm).
    """
    ct = _find_chung_ket()
    if not ct:
        return HttpResponse("Chưa có 'Chung Kết'", status=400)

    columns, rows = _final_columns_and_rows(ct)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{ct.ma}"

    # Header
    ws.append(columns)
    for c in range(1, len(columns)+1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Body
    for r in rows:
        ws.append(r)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    resp = HttpResponse(out.read(),
                        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="export_chungket_{ct.ma}.xlsx"'
    return resp
