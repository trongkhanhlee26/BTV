# after
import csv
import unicodedata  # NEW
import re          # NEW
from io import TextIOWrapper
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render, redirect
from django.db import transaction
from django.core.files.uploadedfile import InMemoryUploadedFile, TemporaryUploadedFile
from openpyxl import load_workbook
from django.db.models import Prefetch
from .models import BaiThi, BaiThiTemplateSection, BaiThiTemplateItem, VongThi  # SAU: thêm import các model template

from .models import ThiSinh, GiamKhao, CuocThi

REQUIRED_COLUMNS = {
    "thisinh": ["maNV", "hoTen", "chiNhanh", "vung", "donVi", "email", "nhom"],
    "giamkhao": ["maNV", "hoTen", "email"],
}

# ===== Alias & chuẩn hóa header =====
def _normalize(s: str) -> str:
    s = unicodedata.normalize("NFKD", s)
    s = s.encode("ascii", "ignore").decode("ascii")  # bỏ dấu
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", "", s)  # bỏ ký tự lạ & khoảng trắng
    return s

HEADER_ALIASES = {
    # maNV
    "manv": "maNV", "manhanvien": "maNV", "manvnv": "maNV", "ma": "maNV", "ma_nv": "maNV",
    # hoTen
    "hoten": "hoTen", "ten": "hoTen", "hovaten": "hoTen", "ho_ten": "hoTen",
    # chiNhanh
    "chinhanh": "chiNhanh", "chi_nhanh": "chiNhanh", "cn": "chiNhanh",
    # vung
    "vung": "vung", "mien": "vung",
    # donVi
    "donvi": "donVi", "don_vi": "donVi", "dv": "donVi", "don": "donVi", "donvichuyendoi": "donVi",
    # email
    "email": "email", "mail": "email", "e-mail": "email",
    # nhom
    "nhom": "nhom", "group": "nhom", "nhomthi": "nhom",
}

def _map_header_list(header, expected_cols):
    """
    Trả về: (canon_order, source_idx)
    canon_order: danh sách tên cột đã được map về canonical (theo thứ tự header gốc)
    source_idx: dict {canonical_name: index_goc}
    """
    canon_order = []
    for h in header:
        key = _normalize(h or "")
        canon = HEADER_ALIASES.get(key)
        canon_order.append(canon or (h or "").strip())

    src_idx = {}
    for i, canon in enumerate(canon_order):
        if canon not in src_idx:
            src_idx[canon] = i

    missing = [c for c in expected_cols if c not in src_idx]
    return canon_order, src_idx, missing

# after
def _read_xlsx(file, expected_cols):
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip() if c is not None else "" for c in rows[0]]

    _, src_idx, missing = _map_header_list(header, expected_cols)
    if missing:
        raise ValueError(f"Thiếu cột: {', '.join(missing)}")

    data = []
    for r in rows[1:]:
        if r is None:
            continue
        row = {}
        for c in expected_cols:
            idx = src_idx[c]
            val = r[idx] if idx < len(r) else None
            row[c] = "" if val is None else str(val).strip()
        data.append(row)
    return data

# after
def _read_csv(file, expected_cols):
    # file là UploadedFile -> cần decode UTF-8
    text_stream = TextIOWrapper(file, encoding="utf-8")
    reader = csv.DictReader(text_stream)
    header = reader.fieldnames or []

    # map header gốc -> canonical
    canon_order, src_idx, missing = _map_header_list(header, expected_cols)
    if missing:
        raise ValueError(f"Thiếu cột: {', '.join(missing)}")

    # Xây map canonical -> tên cột gốc để lấy dữ liệu
    # (ưu tiên đúng chữ canonical nếu đã có, nếu không lấy từ alias)
    canon_to_source = {}
    for i, h in enumerate(header):
        key = _normalize(h or "")
        canon = HEADER_ALIASES.get(key) or (h or "").strip()
        if canon not in canon_to_source:
            canon_to_source[canon] = h

    data = []
    for row in reader:
        out = {}
        for c in expected_cols:
            src = canon_to_source.get(c, c)
            out[c] = (row.get(src, "") or "").strip()
        data.append(out)
    return data
    
@staff_member_required
def import_view(request):
    if request.method == "POST":
        target = request.POST.get("target")  # thisinh | giamkhao
        selected_ma_ct = request.POST.get("maCT")  # NEW
        f = request.FILES.get("file")

        cuocthi_obj = None
        if selected_ma_ct:
            cuocthi_obj = CuocThi.objects.filter(ma=selected_ma_ct).first()

        if target not in REQUIRED_COLUMNS:
            messages.error(request, "Vui lòng chọn loại dữ liệu hợp lệ.")
            return redirect(request.path)
        if not f:
            messages.error(request, "Vui lòng chọn tệp CSV/XLSX.")
            return redirect(request.path)

        expected = REQUIRED_COLUMNS[target]
        try:
            if isinstance(f, (InMemoryUploadedFile, TemporaryUploadedFile)) and f.name.lower().endswith(".xlsx"):
                rows = _read_xlsx(f, expected)
            else:
                rows = _read_csv(f, expected)
        except Exception as e:
            messages.error(request, f"Lỗi đọc tệp: {e}")
            return redirect(request.path)

        created = updated = skipped = 0
        with transaction.atomic():
            if target == "thisinh":
                for r in rows:
                    ma = (r["maNV"] or "").strip()
                    if not ma:
                        skipped += 1
                        continue

                    # Cập nhật chỉ khi trùng cả (maNV, cuộc thi); nếu khác 1 trong 2 → thêm mới
                    # Lưu ý: yêu cầu form đã chọn maCT (cuocthi_obj)
                    lookup = {
                        "maNV": ma,
                        "cuocThi": cuocthi_obj,               # so khớp theo cuộc thi đã chọn
                    }

                    # Nếu model bạn có field maCuocThi (string) để tra cứu nhanh, có thể bổ sung vào defaults:
                    defaults = dict(
                        hoTen=r["hoTen"],
                        chiNhanh=r["chiNhanh"],
                        vung=r["vung"],
                        donVi=r["donVi"],
                        email=r["email"],
                        nhom=r["nhom"],
                        cuocThi=cuocthi_obj,
                    )
                    if hasattr(ThiSinh, "maCuocThi") and cuocthi_obj:
                        defaults["maCuocThi"] = cuocthi_obj.ma

                    obj, is_created = ThiSinh.objects.update_or_create(
                        **lookup,
                        defaults=defaults
                    )

                    created += int(is_created)
                    updated += int(not is_created)

            else:  # giamkhao
                for r in rows:
                    ma = r["maNV"]
                    if not ma:
                        skipped += 1
                        continue
                    obj, is_created = GiamKhao.objects.update_or_create(
                        pk=ma,
                        defaults=dict(
                            hoTen=r["hoTen"],
                            email=r["email"],
                        )
                    )
                    created += int(is_created)
                    updated += int(not is_created)

        messages.success(request, f"Import xong: thêm {created}, cập nhật {updated}, bỏ qua {skipped}.")
        return redirect(request.path)

    return render(
    request,
    "importer/index.html",
    {"cuocthi_list": CuocThi.objects.all().values("ma", "tenCuocThi").order_by("ma")}
)


def organize_view(request):
    if request.method == "POST":
        action = request.POST.get("action")
        # ===== Import MẪU CHẤM cho 1 Bài thi =====
        if action == "config_template_upload":
            btid = request.POST.get("baiThi_id")
            f = request.FILES.get("template_file")
            if not btid or not f:
                messages.error(request, "Thiếu Bài thi hoặc tệp Excel.")
                return redirect(request.path)

            # Lấy bài thi
            try:
                bai_thi = BaiThi.objects.get(pk=btid)
            except BaiThi.DoesNotExist:
                messages.error(request, "Không tìm thấy Bài thi.")
                return redirect(request.path)

            # Đọc Excel bằng openpyxl
            try:
                wb = load_workbook(f, read_only=True, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
            except Exception as e:
                messages.error(request, f"Lỗi đọc Excel: {e}")
                return redirect(request.path)

            if not rows:
                messages.error(request, "Tệp rỗng.")
                return redirect(request.path)

            # Header & alias
            header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
            def idx(*names):
                for name in names:
                    if name in header:
                        return header.index(name)
                return None

            idx_section = idx("section", "mục lớn", "muc lon", "phan", "phần")
            idx_item    = idx("item", "mục nhỏ", "muc nho", "bài", "bai")
            idx_max     = idx("điểm", "diem", "max", "điểm tối đa", "diem toi da")
            idx_note    = idx("note", "ghi chú", "ghi chu")

            if idx_section is None or idx_max is None or (idx_item is None and idx_section is None):
                messages.error(request, "Thiếu cột bắt buộc: Section/Mục lớn, Item/Mục nhỏ (nếu có), Điểm/Max.")
                return redirect(request.path)

            # Xóa template cũ
            BaiThiTemplateSection.objects.filter(baiThi=bai_thi).delete()

            # Parse & Lưu mới theo thứ tự xuất hiện
            section_map = {}   # title -> (section_obj, next_item_stt)
            next_section_stt = 1
            created_items = 0

            for r in rows[1:]:
                if r is None:
                    continue
                sec_title = (r[idx_section] if idx_section is not None and idx_section < len(r) else None)
                item_text = (r[idx_item] if idx_item is not None and idx_item < len(r) else None)
                max_val   = (r[idx_max] if idx_max is not None and idx_max < len(r) else None)
                note_text = (r[idx_note] if idx_note is not None and idx_note < len(r) else None)

                # Bỏ dòng trống hoàn toàn
                if not (sec_title or item_text or max_val or note_text):
                    continue

                sec_title = str(sec_title).strip() if sec_title is not None else ""
                item_text = str(item_text).strip() if item_text is not None else ""
                note_text = str(note_text).strip() if note_text is not None else ""

                # Chuẩn hoá điểm tối đa
                try:
                    max_score = int(float(max_val)) if max_val is not None and str(max_val).strip() != "" else 0
                except Exception:
                    messages.error(request, f"Điểm tối đa không hợp lệ ở dòng có mục: '{item_text or sec_title}'.")
                    return redirect(request.path)

                # Tạo/lấy Section
                if sec_title not in section_map:
                    s = BaiThiTemplateSection.objects.create(
                        baiThi=bai_thi, stt=next_section_stt, title=sec_title or "Mục"
                    )
                    section_map[sec_title] = [s, 1]  # (section_obj, next_item_stt)
                    next_section_stt += 1
                else:
                    s, _ = section_map[sec_title]

                # Có Item (lá) -> tạo item dưới section
                if item_text:
                    s, next_item_stt = section_map[sec_title]
                    BaiThiTemplateItem.objects.create(
                        section=s, stt=next_item_stt, content=item_text, max_score=max_score, note=note_text or None
                    )
                    section_map[sec_title][1] = next_item_stt + 1
                    created_items += 1
                else:
                    # Không có Item => hiểu là "nhánh không có lá": lưu 1 item đặc biệt
                    s, next_item_stt = section_map[sec_title]
                    BaiThiTemplateItem.objects.create(
                        section=s, stt=next_item_stt, content=s.title, max_score=max_score, note=note_text or None
                    )
                    section_map[sec_title][1] = next_item_stt + 1
                    created_items += 1

            messages.success(request, f"Đã cập nhật mẫu chấm cho {bai_thi.tenBaiThi}: {len(section_map)} mục lớn, {created_items} mục nhỏ.")
            return redirect(request.path)

        # (Các action khác như create_ct/create_vt/create_bt/config_time_rules có thể bổ sung sau)
        messages.error(request, "Hành động chưa được hỗ trợ.")
        return redirect(request.path)

    # GET: hiển thị trang tổ chức
    cuoc_this = (
        CuocThi.objects
        .prefetch_related(
            Prefetch("vong_thi", queryset=VongThi.objects.prefetch_related(
                Prefetch("bai_thi", queryset=BaiThi.objects.prefetch_related("time_rules", "template_sections__items"))
            ))
        )
        .order_by("ma")
    )
    return render(request, "organize/index.html", {"cuoc_this": cuoc_this})